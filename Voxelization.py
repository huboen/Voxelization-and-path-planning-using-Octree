from pyOctree_Hu import Octree
import numpy as np
from stl import mesh
import cupy as cp
import time
from numba import cuda
import numba
from Cuda_operator import *
import  openpyxl
import os

class BoundingboxTool:
    def __init__(self,stl_path) -> None:
        self.stl_path = stl_path

    def read_stl(self,file_path=None):
        if file_path == None:
            file_path = self.stl_path
        
        # read stl file
        # mesh_data = o3d.io.read_triangle_mesh(file_path)
        mesh_data = mesh.Mesh.from_file(file_path)
        # obtain the triangles
        # triangles = mesh_data.triangles
        triangles = mesh_data.vectors
        normal_vectors = mesh_data.normals
        
        return triangles, normal_vectors

    @staticmethod
    def boundingBoxes(triangles):
        boundingBoxGroup = []
        for triangle in triangles:
            boundingBoxGroup.append(BoundingboxTool.boundingBox(triangle))
        return boundingBoxGroup
    
    @staticmethod
    def boundingBoxes_gpu(triangles):
        triangles_gpu = cp.asarray(triangles)
        min_bound = cp.min(triangles_gpu, axis=1)
        max_bound = cp.max(triangles_gpu, axis=1)
        bounding_box = [min_bound, max_bound]
        bounding_boxes = cp.stack(bounding_box, axis=-1)
        return bounding_boxes.get()

    @staticmethod
    def boundingBox(triangle):
        min_bound = np.min(triangle, axis=0)
        max_bound = np.max(triangle, axis=0)
        bounding_box = [min_bound, max_bound]
        return bounding_box
    
    def maxBoundingBox(self):
        triangles,_=self.read_stl(self.stl_path)
        vertices= triangles.reshape(-1,3)
        min_bound = np.min(vertices,axis=0)
        max_bound = np.max(vertices,axis=0)
        max_bounding_box = [min_bound,max_bound]
        return max_bounding_box

    def minBoundingBox(self):
        triangles,_ = self.read_stl()
        bounding_boxes = np.array(BoundingboxTool.boundingBoxes_gpu(triangles))
        size = np.min(np.max(bounding_boxes[:,:,1]-bounding_boxes[:,:,0],axis=1))
        return size
    
    staticmethod
    def get_tri_info():
        pass
        
        
 

    

class OctreeOperator(Octree):
    def __init__(self, boundingbox) -> None:
        super().__init__(boundingbox)

    # def boundingBoxIntersect(self, node, targetBoundingBox):
    # # Check if two bounding boxes intersect
    #     miniNode, maxNode = np.array(node.center) - node.size / 2, np.array(node.center) + node.size / 2 
    #     miniBbox, maxBbox= targetBoundingBox[0], targetBoundingBox[1]

    #     return not (maxNode[0] < miniBbox[0] or miniNode[0] > maxBbox[0] or
    #                 maxNode[1] < miniBbox[1] or miniNode[1] > maxBbox[1] or
    #                 maxNode[2] < miniBbox[2] or miniNode[2] > maxBbox[2] or
    #                 maxBbox[0] < miniNode[0] or miniBbox[0] > maxNode[0] or
    #                 maxBbox[1] < miniNode[1] or miniBbox[1] > maxNode[1] or
    #                 maxBbox[2] < miniNode[2] or miniBbox[2] > maxNode[2])
    @staticmethod
    @cuda.jit
    def boundingBoxIntersect_cuda(centers, sizes, targetBoundingBoxes, results,output_test = None):
        thread= cuda.grid(1)
        if thread < targetBoundingBoxes.shape[0]:
            # miniNode, maxNode = centers - sizes / 2, centers + sizes / 2 
                # 先获取 targetBoundingBoxes 切片
                if output_test is not None:
                    output_test[thread] = thread
                bbox_slice = targetBoundingBoxes[thread]
                
                # 然后再获取 [0, :] 切片
                miniBbox = cuda.local.array(3, dtype=numba.float32)
                maxBbox = cuda.local.array(3, dtype=numba.float32)
                miniBbox, maxBbox = bbox_slice[0], bbox_slice[1]
                for node_number in range(len(centers)):
                    miniNode = cuda.local.array(3, dtype=numba.float32)
                    maxNode = cuda.local.array(3, dtype=numba.float32)
                    miniNode[0] = centers[node_number, 0] - sizes[node_number, 0]/2
                    miniNode[1] = centers[node_number, 1] - sizes[node_number, 1]/2
                    miniNode[2] = centers[node_number, 2] - sizes[node_number, 2]/2
                    maxNode[0] = centers[node_number, 0] + sizes[node_number, 0]/2
                    maxNode[1] = centers[node_number, 1] + sizes[node_number, 1]/2
                    maxNode[2] = centers[node_number, 2] + sizes[node_number, 2]/2

                    results[thread,node_number]= not    (maxNode[0] < miniBbox[0] or miniNode[0] > maxBbox[0] or
                                                        maxNode[1] < miniBbox[1] or miniNode[1] > maxBbox[1] or
                                                        maxNode[2] < miniBbox[2] or miniNode[2] > maxBbox[2] )
 



    #find bounding Nodes for one layer with cuda                
    def findBoundingNodesOnce_cuda(self,target_bounding_boxes,intersected_nodes = None):
        if intersected_nodes is None:
            centers, sizes = OctreeOperator.maxtrixOperator(self.all_leaf_nodes(self.root.children[0]))
        else:
            centers, sizes = OctreeOperator.maxtrixOperator(intersected_nodes)
            
        
        centers_gpu = cuda.to_device(np.array(centers))
        sizes_gpu = cuda.to_device(np.array(sizes))
        target_bounding_box_gpu = cuda.to_device(np.ascontiguousarray(target_bounding_boxes))
        result_gpu = cuda.device_array((target_bounding_box_gpu.shape[0],centers_gpu.shape[0]),dtype=np.int8)
        output_test_gpu = cuda.device_array(target_bounding_box_gpu.shape[0],dtype=np.int8)
        # 设置线程和块大小
        threads_per_block = 256
        blocks_per_grid = (target_bounding_box_gpu.shape[0] + threads_per_block - 1) // threads_per_block
        
        # 调用 CUDA 核函数
        OctreeOperator.boundingBoxIntersect_cuda[blocks_per_grid, threads_per_block](centers_gpu, sizes_gpu, target_bounding_box_gpu, result_gpu, output_test_gpu)
        
        # 从 GPU 获取结果
        # cuda.synchronize()
        # print("at 1")
        begin = time.time()
        result_2d= result_gpu.copy_to_host()
        end = time.time()
        duration = end - begin
        # print("at 2")
        print("amount:",centers_gpu.shape[0],"get result time",duration, "storage:",result_2d.nbytes/(1024*1024*1024))
        result = OctreeOperator.reduce_size(result_2d)
        # for i in output_test_gpu.copy_to_host():
        #     print(i)
        return result
    
    # find all BoundingNodes of target depth with cuda
    def findBoundingNodesAll_cuda(self, target_depth, target_bounding_boxes,intersected_nodes=None):
        if intersected_nodes is None:
            intersected_tree_nodes = self.root.children
        else:
            intersected_tree_nodes = intersected_nodes
        max_depth = intersected_tree_nodes[0].depth
        results = np.ones(len(intersected_tree_nodes), dtype=np.int8)
        while max_depth < target_depth:
            for i in range(results.size):
                if results[i]:
                    OctreeOperator.extend(intersected_tree_nodes[i],intersected_tree_nodes[i].depth+1)
            intersected_tree_nodes = self.intersected_node_update(results,intersected_tree_nodes)
            # print("go in")
            results = self.findBoundingNodesOnce_cuda(target_bounding_boxes=target_bounding_boxes,intersected_nodes=intersected_tree_nodes)
            # print("go out")
            
            
            max_depth = intersected_tree_nodes[0].depth
            # self.all_leaf_nodes()
        # intersected_tree_nodes = np.where(results==True,intersected_tree_nodes,0)
        other_nodes = np.delete(intersected_tree_nodes, np.where(results==True) )
        intersected_tree_nodes = np.delete(intersected_tree_nodes, np.where(results==False) )
        
            # print(max_depth)
        # print("end")
        return intersected_tree_nodes,other_nodes

    
    def __update_octree__(self, results, target_depth):
        for label, node in enumerate(results,self.leafnodes):
            depth = node.depth()
            if label and depth < target_depth:
                self.extend(node, depth+1)
        self.all_leaf_nodes()
    @staticmethod
    def maxtrixOperator(nodes):
        centers = []
        sizes = []
        for node in nodes:
            centers.append(node.center)
            sizes.append(np.full(3, node.size, dtype=np.float32))
        return np.array(centers), np.array(sizes)
    @staticmethod
    def reduce_size(arr):
    # 沿着第一维度（轴0）应用any函数
        result = np.any(arr, axis=0)
        return result
    def intersected_node_update(self,results,intersected_nodes_old):
        intersected_tree_nodes_new = []
        for i in range(results.size):
            if results[i]:
                if len(intersected_nodes_old[i].children) :
                    intersected_tree_nodes_new.extend(intersected_nodes_old[i].children)
                else:
                    intersected_tree_nodes_new.extend(intersected_nodes_old[i])
        return intersected_tree_nodes_new
    
    @staticmethod
    def transferNode2box(nodes):
        boundingBoxes = []
        for node in nodes:
            min_bound = np.array(node.center)- node.size
            max_bound = np.array(node.center)+ node.size
            boundingBoxes.append([min_bound,max_bound])
        return boundingBoxes
    
    @staticmethod
    def write_to_excel(nodes,name):
        if os.path.exists(name):
            print(f"File {name} already exists. Choose a different file name.")
            return
        workbook = openpyxl.Workbook()
        for i in range(0, len(nodes), 1048575):  # 每个文件最多写入1048575行
            sheet = workbook.create_sheet(title=f'Sheet_{i//1048575 + 1}')
            sheet['A1'] = 'center_x'
            sheet['B1'] = 'center_Y'
            sheet['C1'] = 'center_Z'
            sheet['D1'] = 'size'
            sheet['E1'] = "depth"

            for j in range(i, min(i+1048575, len(nodes))):
                row = j - i + 2
                sheet.cell(row=row, column=1, value=nodes[j].center[0])
                sheet.cell(row=row, column=2, value=nodes[j].center[1])
                sheet.cell(row=row, column=3, value=nodes[j].center[2])
                sheet.cell(row=row, column=4, value=nodes[j].size)
                sheet.cell(row=row, column=5, value=nodes[j].depth)
                # path = "B:\\Master arbeit\\" + name
    # 构建完整的文件路径
        workbook.save(name)
        print("created the node_data")

    @staticmethod
    def transferNode2box(nodes):
        boundingBoxes = []
        for node in nodes:
            min_bound = np.array(node.center)- node.size
            max_bound = np.array(node.center)+ node.size
            boundingBoxes.append([min_bound,max_bound])
        return boundingBoxes
    
class separatingAxis:
    def __init__(self) -> None:
        pass
    
    @staticmethod
    def collision_3d_triangle_box(triangle, tri_normal, node):
        # tri_normal /= np.linalg.norm(tri_normal)

        # # 分离轴1: 三角形法向量
        if not separatingAxis.overlap_on_axis(triangle, tri_normal, node):
            return False

        # # 分离轴2-4: 三角形的边
        tri_edges = [triangle[1] - triangle[0], triangle[2] - triangle[1], triangle[0] - triangle[2]]
        for axis in tri_edges:
            if not separatingAxis.overlap_on_axis(triangle, axis, node):
                return False

        # # 分离轴5-7: 立方体的边
        node_edges = [[1, 0, 0], [0, 1, 0], [0, 0, 1]]
        for axis in node_edges:
            if not separatingAxis.overlap_on_axis(triangle, axis, node):
                return False
            
        for tri_edge in tri_edges:
            for node_edge in node_edges:
                axis = np.cross(tri_edge,node_edge)
                if not separatingAxis.overlap_on_axis(triangle, axis, node):
                    return False
        # 如果所有轴上的投影都有交集，则相交
        return True
    
    @staticmethod
    def overlap_on_axis(triangle, axis, node):
        tri_proj = [np.dot(tri_vertex, axis) for tri_vertex in triangle]
        node_vertexes = separatingAxis.node_vertexes_from_center_and_size(node.center, node.size)
        node_proj = [np.dot(node_vertex, axis) for node_vertex in node_vertexes]

        return not (max(tri_proj) < min(node_proj) or min(tri_proj) > max(node_proj))
    
    @staticmethod
    def node_vertexes_from_center_and_size(center, size):
        return [
            [center[0] - size / 2, center[1] - size / 2, center[2] - size / 2],
            [center[0] - size / 2, center[1] - size / 2, center[2] + size / 2],
            [center[0] - size / 2, center[1] + size / 2, center[2] - size / 2],
            [center[0] - size / 2, center[1] + size / 2, center[2] + size / 2],
            [center[0] + size / 2, center[1] - size / 2, center[2] - size / 2],
            [center[0] + size / 2, center[1] - size / 2, center[2] + size / 2],
            [center[0] + size / 2, center[1] + size / 2, center[2] - size / 2],
            [center[0] + size / 2, center[1] + size / 2, center[2] + size / 2]
        ]

        
    @staticmethod
    @cuda.jit
    def collision_3d_triangle_box_cuda(triangles_vertex,triangles_normals,nodes_center,nodes_vertex,node_half_size,node_edges, results):
        
        tri_thread,node_thread= cuda.grid(2)
        if tri_thread < triangles_normals.shape[0] and node_thread < nodes_center.shape[0]:
            d_triangle_edges = cuda.local.array((3,3),dtype=np.float32)
            tri_normal = triangles_normals[tri_thread]
            d_node_center = nodes_center[node_thread]
            d_node_vertex = nodes_vertex[node_thread]
            d_tri_vertex = triangles_vertex[tri_thread]
            d_node_edges = node_edges
            triangle_edge(d_tri_vertex,d_triangle_edges)
            tri_proj = cuda.local.array(3,dtype=np.float32)
            node_proj = cuda.local.array(8,dtype=np.float32)
            if not overlap_on_axis(d_tri_vertex,d_node_vertex,tri_normal,tri_proj,node_proj):
                return
            for axis in d_triangle_edges:
                if not overlap_on_axis(d_tri_vertex,d_node_vertex,axis,tri_proj,node_proj):
                    return
            for axis in d_node_edges:
                if not overlap_on_axis(d_tri_vertex,d_node_vertex,axis,tri_proj,node_proj):
                    return
            for d_tri_edge in d_triangle_edges:
                for d_node_edge in d_node_edges:
                    axis = cuda.local.array(3,dtype=np.float32)
                    cross(d_tri_edge,d_node_edge,axis)
                    if not overlap_on_axis(d_tri_vertex,d_node_vertex,axis,tri_proj,node_proj):
                        return
            results[tri_thread,node_thread] = True
    @staticmethod
    def node_info(nodes):
        vertexes = []
        centers = []
        sizes = []
        for i in range(len(nodes)):
            node_center = nodes[i].center
            size = nodes[i].size
            centers.append(node_center)
            sizes.append([size/2,size/2,size/2])
            vertexes.append([[node_center[0]- size/2,node_center[1] - size/2,node_center[2] - size/2],
                            [node_center[0]- size/2,node_center[1]  - size/2,node_center[2]  + size/2],
                            [node_center[0]- size/2,node_center[1]  + size/2,node_center[2]  - size/2],
                            [node_center[0]- size/2,node_center[1]  + size/2,node_center[2]  + size/2],
                            [node_center[0]+ size/2,node_center[1]  - size/2,node_center[2]  - size/2],
                            [node_center[0]+ size/2,node_center[1]  - size/2,node_center[2]  + size/2],
                            [node_center[0]+ size/2,node_center[1]  + size/2,node_center[2]  - size/2],
                            [node_center[0]+ size/2,node_center[1]  + size/2,node_center[2]  + size/2]])
        vertexes = np.array(vertexes,dtype=np.float32)
        centers = np.array(centers,dtype=np.float32)
        sizes = np.array(sizes,dtype=np.float32)
        return vertexes,centers, sizes
    
    @staticmethod
    def separatingAxis_calculation(triangles_vertexes,triangles_normals,nodes,GPU_memory = None):
        if GPU_memory is None:
            GPU_memory = 8 # The GPU I use is 8 GB
        node_vertexes,node_centers,node_half_sizes = separatingAxis.node_info(nodes)
        node_1 = node_vertexes[0]
        node_vectors = [[0,0,1],[0,1,0],[1,0,0]]
        result = []
        threads_per_block = (16,16)
        blocks_per_grid_x = (triangles_normals.shape[0]+ threads_per_block[0]- 1) // threads_per_block[0]
        blocks_per_grid_y = (len(nodes)+ threads_per_block[1] - 1) // threads_per_block[1]
        blocks_per_grid = (blocks_per_grid_x,blocks_per_grid_y)
        d_results = cuda.device_array((triangles_vertexes.shape[0],node_centers.shape[0]),dtype=np.int8)
        d_tri_vertexes = cuda.to_device(np.ascontiguousarray(triangles_vertexes))
        d_tri_normals = cuda.to_device(np.ascontiguousarray(triangles_normals))
        d_node_centers = cuda.to_device(np.ascontiguousarray(node_centers))
        d_node_vertexes = cuda.to_device(np.ascontiguousarray(node_vertexes))
        d_node_sizes = cuda.to_device(np.ascontiguousarray(node_half_sizes))
        d_node_vectors = cuda.to_device(np.ascontiguousarray(node_vectors))
        separatingAxis.collision_3d_triangle_box_cuda[blocks_per_grid,threads_per_block](d_tri_vertexes,d_tri_normals,d_node_centers,d_node_vertexes ,d_node_sizes,d_node_vectors,d_results)
        begin = time.time()
        result=d_results.copy_to_host()
        end = time.time()
        print("nodes_amount",len(nodes),"time",end-begin,"storage",result.nbytes/(1024*1024*1024))
        result = OctreeOperator.reduce_size(result)
        intersected_nodes=np.delete(nodes,np.where(result==False))
        return intersected_nodes 
    
    @staticmethod
    @cuda.jit
    def moeller_method(ray_origins,ray_vectors,triangles,d_epsilon,results):
        i,j = cuda.grid(2)
        n_rays,n_triangles = ray_origins.shape[0],triangles.shape[0]

        if i <n_triangles and j < n_rays:
            ray_origin = ray_origins[j]
            ray_vector = ray_vectors
            triangle = triangles[i]
            edge1 = cuda.local.array(3,dtype=np.float32)
            edge2 = cuda.local.array(3,dtype=np.float32)
            sub(triangle[1],triangle[0],edge1)
            sub(triangle[2],triangle[0],edge2)
            epsilon = d_epsilon[0]
            h = cuda.local.array(3,dtype=np.float32)
            cross(ray_vector,edge2,h)
            a = dot(edge1,h)
            if  a < epsilon and a > -epsilon :
                results[i,j] = False
                # return
            else:
                f = 1/a
                s = cuda.local.array(3,dtype=np.float32)
                sub(ray_origin,triangle[0],s)
                u = f*dot(s,h)
                if 0 <=u<=1:
                    q = cuda.local.array(3,dtype=np.float32)
                    cross(s,edge1,q)
                    v = f*dot(ray_vector,q)
                    if 0 <=v<=1 and u+v<=1:
                        t = f *dot(edge2,q)
                        if t>epsilon:
                            results[i,j] = True
                            return
            results[i,j]=False

    @staticmethod
    def ray_triangle_intersection(triangles,nodes):
        _,centers,_ = separatingAxis.node_info(nodes)
        threads_per_block = (16,16)
        blocks_per_grid_x = (triangles.shape[0]+ threads_per_block[0]- 1) // threads_per_block[0]
        blocks_per_grid_y = (len(nodes)+ threads_per_block[1] - 1) // threads_per_block[1]
        blocks_per_grid = (blocks_per_grid_x,blocks_per_grid_y)
        d_centers = cuda.to_device(np.array(centers,dtype=np.float32))
        d_triangles = cuda.to_device(np.array(triangles,dtype=np.float32))
        d_vector = cuda.to_device(np.array([1,0,0],dtype=np.float32))
        d_results = cuda.device_array((triangles.shape[0],len(nodes)),dtype=np.int8)
        d_epsilion = cuda.to_device(np.array([1e-8],dtype=np.float32))
        separatingAxis.moeller_method[blocks_per_grid,threads_per_block](d_centers,d_vector,d_triangles,d_epsilion,d_results)
        inner_nodes = []
        results = d_results.copy_to_host()
        begin = time.time()
        results = np.sum(results,axis = 0)
        inner_nodes = np.delete(nodes,np.where(results%2==0))
        end = time.time()
        print("time",end-begin)
        return inner_nodes

class Voxelization_exe():
    def __init__(self,data_path) -> None:
        self.data_path = data_path

    
    def run(self,targetDepth,excel=False,vis=False):
        #initilize important parameters
        voxl = BoundingboxTool(self.data_path)
        triangles,normalVectors = voxl.read_stl()
        normalVectors = np.array(normalVectors,dtype=np.float32)
        bounding_boxes = np.array(voxl.boundingBoxes_gpu(triangles))
        maxBoundingBox = voxl.maxBoundingBox()
        minBoundingBox = voxl.minBoundingBox()
        octreeTest = OctreeOperator(maxBoundingBox)
        targetboundingbox =bounding_boxes[0].T
        targetboundingboxes =np.transpose(bounding_boxes, (0, 2, 1))
        target_depth= targetDepth

        #find nodes intersected with triangle boundingboxes
        start = time.time()
        intersected_nodes,_ = octreeTest.findBoundingNodesAll_cuda(target_depth=5,target_bounding_boxes=targetboundingboxes,intersected_nodes=None)
        depth = intersected_nodes[0].depth
        GPU_memory = 8
        other_nodes = []
        while depth<target_depth:
            intersected_nodes_split=np.array_split(intersected_nodes, np.ceil(len(intersected_nodes) /5000))
            intersected_nodes = []
            for group in intersected_nodes_split:
                intersec,other= octreeTest.findBoundingNodesAll_cuda(target_depth=depth+1,target_bounding_boxes=targetboundingboxes,intersected_nodes=group)
                intersected_nodes.extend(intersec)
                other_nodes.extend(other)
            depth = intersected_nodes[0].depth

        #find nodes intersected with triangles    
        size1 = len(intersected_nodes)
        intersected_nodes_split= np.array_split(intersected_nodes,np.ceil(len(intersected_nodes)/40000))
        intersected_nodes=[]
        for group in intersected_nodes_split:
            intersected_nodes.extend(separatingAxis.separatingAxis_calculation(triangles,normalVectors,group))
        size2 = len(intersected_nodes)
        end = time.time()
        duration = end - start

        #update the octree leafnodes
        octreeTest.all_leaf_nodes()
        resolution = octreeTest.root.size*(0.5**(target_depth))

        # find the nodes inside the object
        other_nodes_splited = np.array_split(other_nodes,np.ceil(len(other_nodes)/40000))
        inner_nodes=[]
        for group in other_nodes_splited:
            inner_nodes.extend(separatingAxis.ray_triangle_intersection(triangles,group))
        inner_nodes = [node for node in inner_nodes if node.depth == target_depth]
        print("Computation time", duration)
        print("resolution",resolution)
        print("how many leafnodes",len(octreeTest.leafnodes))
        print("how many intersected nodes",len(intersected_nodes))
        print("how many other leafnodes",len(other_nodes))
        print("change of intersected nodes", size2-size1)
        print("amount of inner nodes", len(inner_nodes))

        #used for visulization
        node_boxes = OctreeOperator.transferNode2box(inner_nodes)


        # generate excel file for data
        if excel:
            name1="inner_nodes" + str(target_depth) +".xlsx"
            name2="node_data" + str(target_depth) +".xlsx"
            current_dir = os.path.dirname(os.path.abspath(__file__))
            file_path1 = os.path.join(current_dir,"node_data", name1)
            file_path2 = os.path.join(current_dir,"node_data", name2)
            if len(inner_nodes):
                octreeTest.write_to_excel(inner_nodes,name=file_path1)
            if len(intersected_nodes):
                octreeTest.write_to_excel(intersected_nodes,name=file_path2)
        # visilize but not for large data
        if vis:
            octreeTest.visualize(stl_path=self.data_path,boundingboxes=node_boxes,octree=False)

if __name__ == "__main__":
    pass